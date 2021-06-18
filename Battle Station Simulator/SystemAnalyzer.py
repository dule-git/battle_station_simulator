import numpy as np
import matplotlib.pyplot as plt
import math
import openpyxl

class SystemAnalyzer:

    def __init__(self, system):
        self.system = system
        self.theta_azimuth = np.array([])
        self.omega_azimuth = np.array([])
        self.alpha_azimuth = np.array([])

        self.theta_elevation = np.array([])
        self.omega_elevation = np.array([])
        self.alpha_elevation = np.array([])

        self.distance = np.array([])

    def add_snapshot(self):
        if (self.system.distance != 0):
            self.theta_azimuth = np.append(self.theta_azimuth, self.system.azimuth)
            self.omega_azimuth = np.append(self.omega_azimuth, math.degrees(self.system.target.curr_velocity_xy / self.system.distance_xy))
            self.alpha_azimuth = np.append(self.alpha_azimuth, math.degrees(math.pow(self.system.target.curr_velocity_xy, 2) / self.system.distance_xy))

            self.theta_elevation = np.append(self.theta_elevation, self.system.elevation)
            self.omega_elevation = np.append(self.omega_elevation, math.degrees(self.system.target.curr_velocity_z / self.system.distance))
            self.alpha_elevation = np.append(self.alpha_elevation, math.degrees(math.pow(self.system.target.curr_velocity_z, 2) / self.system.distance))

            self.distance = np.append(self.distance, self.system.distance)
        else:
            self.theta_azimuth = np.append(self.theta_azimuth, self.theta_azimuth[-1])
            self.omega_azimuth = np.append(self.omega_azimuth, self.omega_azimuth[-1])
            self.alpha_azimuth = np.append(self.alpha_azimuth, self.alpha_azimuth[-1])

            self.theta_elevation = np.append(self.theta_elevation, self.theta_elevation[-1])
            self.omega_elevation = np.append(self.omega_elevation, self.omega_elevation[-1])
            self.alpha_elevation = np.append(self.alpha_elevation, self.alpha_elevation[-1])

            self.distance = np.append(self.distance, self.system.distance)

    def show_analysis(self, time_step):
        num_of_frames = len(self.system.target.animation_x_coords)
        time = np.arange(0, num_of_frames * time_step, time_step)
        
        theta_azimuth_degrees = np.array([])
        theta_elevation_degrees = np.array([])
        for azimuth, elevation in zip(self.theta_azimuth, self.theta_elevation):
            theta_azimuth_degrees = np.append(theta_azimuth_degrees, math.degrees(azimuth))
            theta_elevation_degrees = np.append(theta_elevation_degrees, math.degrees(elevation))

        fig, axs = plt.subplots(2, 3)

        # azimuth analysis
        axs[0][0].set_title('Azimuth: Theta(time)')
        axs[0][0].set_xlabel('Time[s]')
        axs[0][0].set_ylabel('Theta[degrees]')
        axs[0][0].set_ylim([-200, 200])
        axs[0][0].plot(time, theta_azimuth_degrees, color='blue')

        axs[0][1].set_title('Azimuth: Velocity(time)')
        axs[0][1].set_xlabel('Time[s]')
        axs[0][1].set_ylabel('Velocity[degrees/s]')
        axs[0][1].plot(time, self.omega_azimuth, color='green')

        axs[0][2].set_title('Azimuth: Acceleration(time)')
        axs[0][2].set_xlabel('Time[s]')
        axs[0][2].set_ylabel('Acceleration[degrees/s^2]')
        axs[0][2].plot(time, self.alpha_azimuth, color='red')

        # elevation analysis
        axs[1][0].set_title('Elevation: Theta(time)')
        axs[1][0].set_xlabel('Time[s]')
        axs[1][0].set_ylabel('Theta[degrees]')
        axs[1][0].plot(time, theta_elevation_degrees, color='blue')

        axs[1][1].set_title('Elevation: Velocity(time)')
        axs[1][1].set_xlabel('Time[s]')
        axs[1][1].set_ylabel('Velocity[degrees/s]')
        axs[1][1].plot(time, self.omega_elevation, color='green')

        axs[1][2].set_title('Elevation: Acceleration(time)')
        axs[1][2].set_xlabel('Time[s]')
        axs[1][2].set_ylabel('Acceleration[degrees/s^2]')
        axs[1][2].plot(time, self.alpha_elevation, color='red')

        plt.show()
    
    def export_to_excel(self, kml_file, time_step, path):
        num_of_frames = len(self.system.target.animation_x_coords)
        time = np.arange(0, num_of_frames * time_step, time_step)

        theta_azimuth_degrees = np.array([])
        theta_elevation_degrees = np.array([])
        for azimuth, elevation in zip(self.theta_azimuth, self.theta_elevation):
            theta_azimuth_degrees = np.append(theta_azimuth_degrees, math.degrees(azimuth))
            theta_elevation_degrees = np.append(theta_elevation_degrees, math.degrees(elevation))

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.cell(row = 2, column = 5).value = 'Time [s]'
        sheet.cell(row = 2, column = 6).value = 'Distance [m]'
        sheet.cell(row = 2, column = 7).value = 'Theta [\u00b0]'
        sheet.cell(row = 2, column = 8).value = 'Omega [\u00b0/s]'
        sheet.cell(row = 2, column = 9).value = 'Alpha [\u00b0/s^2]'
        sheet.cell(row = 2, column = 10).value = 'Theta [\u00b0]'
        sheet.cell(row = 2, column = 11).value = 'Omega [\u00b0/s]'
        sheet.cell(row = 2, column = 12).value = 'Alpha [\u00b0/s^2]'
        for cnt, t, theta_azimuth, omega_azimuth, alpha_azimuth, theta_elevation, omega_elevation, alpha_elevation, distance in zip(range(1, len(time) + 1), time, theta_azimuth_degrees, self.omega_azimuth, self.alpha_azimuth, theta_elevation_degrees, self.omega_elevation, self.alpha_elevation, self.distance):
            row_num = cnt + 2
            sheet.cell(row = row_num, column = 5).value = t
            sheet.cell(row = row_num, column = 6).value = distance
            sheet.cell(row = row_num, column = 7).value = theta_azimuth
            sheet.cell(row = row_num, column = 8).value = omega_azimuth
            sheet.cell(row = row_num, column = 9).value = alpha_azimuth
            sheet.cell(row = row_num, column = 10).value = theta_elevation
            sheet.cell(row = row_num, column = 11).value = omega_elevation
            sheet.cell(row = row_num, column = 12).value = alpha_elevation
        
        excel_file_animation = kml_file.replace('files/import/', '')
        excel_file_animation = excel_file_animation.replace('.kml', '')
        excel_file_animation = excel_file_animation + ' Excel.xlsx'
        workbook.save(path + excel_file_animation)
        print('Excel file exported as: ' + path + excel_file_animation)
